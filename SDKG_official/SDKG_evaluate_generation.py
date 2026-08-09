from __future__ import annotations

import argparse
import csv
import difflib
import json
import math
import re
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
QUERY_FILE = BASE_DIR / "data" / "queries_50_0519.xlsx"
DEFAULT_BERT_MODEL = "shibing624/text2vec-base-chinese"
DEFAULT_OUTPUT_DIR = BASE_DIR / "evaluation_outputs"


def load_query_sheet(path: Path) -> tuple[dict[int, dict], list[int]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: header.index(name) for name in header}

    queries: dict[int, dict] = {}
    ordered_ids: list[int] = []
    current_category = ""

    for row in rows:
        query_id = row[idx["query_id"]]
        query_text = row[idx["模擬輸入內容"]]
        if not query_id or not query_text:
            continue

        raw_category = row[idx["類別"]]
        if raw_category not in (None, ""):
            current_category = str(raw_category)

        qid = int(query_id)
        ordered_ids.append(qid)
        queries[qid] = {
            "query_id": qid,
            "category": current_category,
            "query_text": str(query_text),
            "ground_truth_text": str(row[idx["Ground Truth-人工起訴書"]] or ""),
            "gpt_baseline_text": str(row[idx["gpt-4o-mini"]] or ""),
        }

    return queries, ordered_ids


def iter_jsonl(path: Path) -> Iterable[dict]:
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                yield json.loads(line)


def normalize_text(text: str) -> str:
    text = text or ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\s+", "", text)
    return text


def char_tokens(text: str) -> list[str]:
    return list(normalize_text(text))


def ngram_counts(tokens: list[str], n: int) -> dict[tuple[str, ...], int]:
    counts: dict[tuple[str, ...], int] = defaultdict(int)
    if len(tokens) < n or n <= 0:
        return counts
    for i in range(len(tokens) - n + 1):
        counts[tuple(tokens[i:i + n])] += 1
    return counts


def bleu4_score(candidate: str, reference: str) -> float:
    cand = char_tokens(candidate)
    ref = char_tokens(reference)
    if not cand or not ref:
        return 0.0

    precisions = []
    for n in range(1, 5):
        cand_counts = ngram_counts(cand, n)
        ref_counts = ngram_counts(ref, n)
        total = sum(cand_counts.values())
        if total == 0:
            precisions.append(0.0)
            continue
        overlap = 0
        for gram, count in cand_counts.items():
            overlap += min(count, ref_counts.get(gram, 0))
        precisions.append((overlap + 1.0) / (total + 1.0))

    if min(precisions) <= 0:
        geo_mean = 0.0
    else:
        geo_mean = math.exp(sum(math.log(p) for p in precisions) / 4.0)

    cand_len = len(cand)
    ref_len = len(ref)
    brevity_penalty = 1.0 if cand_len > ref_len else math.exp(1.0 - (ref_len / max(cand_len, 1)))
    return brevity_penalty * geo_mean


def lcs_length(left: list[str], right: list[str]) -> int:
    if not left or not right:
        return 0
    prev = [0] * (len(right) + 1)
    for i in range(1, len(left) + 1):
        curr = [0] * (len(right) + 1)
        for j in range(1, len(right) + 1):
            if left[i - 1] == right[j - 1]:
                curr[j] = prev[j - 1] + 1
            else:
                curr[j] = max(prev[j], curr[j - 1])
        prev = curr
    return prev[-1]


def rouge_l_f1(candidate: str, reference: str) -> float:
    cand = char_tokens(candidate)
    ref = char_tokens(reference)
    if not cand or not ref:
        return 0.0
    # Exact character-level LCS is too slow for 1000 long legal documents.
    # SequenceMatcher gives a stable overlap proxy that is practical here.
    return difflib.SequenceMatcher(a=cand, b=ref, autojunk=False).ratio()


class BertScoreProxy:
    def __init__(self, model_name: str = DEFAULT_BERT_MODEL) -> None:
        import torch
        from transformers import AutoModel, AutoTokenizer

        self.torch = torch
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        self.use_fallback = False

        try:
            self.tokenizer = AutoTokenizer.from_pretrained(model_name, local_files_only=True)
            self.model = AutoModel.from_pretrained(model_name, local_files_only=True).to(self.device)
            self.model.eval()
        except Exception:
            self.use_fallback = True
            self.tokenizer = None
            self.model = None

    def _token_embeddings(self, text: str) -> torch.Tensor:
        normalized = normalize_text(text)
        if not normalized:
            return torch.empty((0, 1), device=self.device)
        encoded = self.tokenizer(
            normalized,
            return_tensors="pt",
            truncation=True,
            max_length=512,
        )
        encoded = {key: value.to(self.device) for key, value in encoded.items()}
        with self.torch.no_grad():
            outputs = self.model(**encoded)
        hidden = outputs.last_hidden_state[0]
        mask = encoded["attention_mask"][0].bool()
        token_ids = encoded["input_ids"][0]
        special_mask = self.torch.zeros_like(mask, dtype=self.torch.bool)
        for token_id in self.tokenizer.all_special_ids:
            special_mask |= token_ids.eq(token_id)
        keep = mask & (~special_mask)
        return hidden[keep]

    def score(self, candidate: str, reference: str) -> float:
        if self.use_fallback:
            return bertscore_fallback(candidate, reference)
        cand = self._token_embeddings(candidate)
        ref = self._token_embeddings(reference)
        if cand.numel() == 0 or ref.numel() == 0:
            return 0.0

        cand = self.torch.nn.functional.normalize(cand, p=2, dim=1)
        ref = self.torch.nn.functional.normalize(ref, p=2, dim=1)
        sim = cand @ ref.T
        precision = sim.max(dim=1).values.mean()
        recall = sim.max(dim=0).values.mean()
        denom = precision + recall
        if float(denom) == 0.0:
            return 0.0
        f1 = 2.0 * precision * recall / denom
        return float(f1.item())


def bertscore_fallback(candidate: str, reference: str) -> float:
    cand = char_tokens(candidate)
    ref = char_tokens(reference)
    if not cand or not ref:
        return 0.0
    cand_set = set(cand)
    ref_set = set(ref)
    overlap = len(cand_set & ref_set)
    precision = overlap / len(cand_set) if cand_set else 0.0
    recall = overlap / len(ref_set) if ref_set else 0.0
    if precision + recall == 0:
        return 0.0
    return (2.0 * precision * recall) / (precision + recall)


def parse_sdkg_records(paths: list[Path]) -> list[dict]:
    rows = []
    for path in paths:
        for record in iter_jsonl(path):
            if record.get("run_status") == "error":
                continue
            rows.append({
                "source_file": str(path),
                "family": "DKG",
                "system_label": str(record.get("exp_name") or record.get("exp_id") or "DKG"),
                "query_id": int(record["query_id"]),
                "top_k": int(record.get("top_k", 0) or 0),
                "generated_text": str(record.get("generated_text") or ""),
            })
    return rows


def parse_taarn_records(paths: list[Path]) -> list[dict]:
    rows = []
    for path in paths:
        for record in iter_jsonl(path):
            rows.append({
                "source_file": str(path),
                "family": "TAARN",
                "system_label": "TAARN",
                "query_id": int(record["query_id"]),
                "top_k": int(record.get("top_k", 0) or 0),
                "generated_text": str(record.get("generated_text") or ""),
            })
    return rows


def build_gpt_rows(queries: dict[int, dict], top_k_values: list[int]) -> list[dict]:
    rows = []
    for top_k in top_k_values:
        for query in queries.values():
            rows.append({
                "source_file": str(QUERY_FILE),
                "family": "GPT",
                "system_label": "gpt-4o-mini",
                "query_id": int(query["query_id"]),
                "top_k": int(top_k),
                "generated_text": str(query["gpt_baseline_text"] or ""),
            })
    return rows


def evaluate_rows(
    rows: list[dict],
    queries: dict[int, dict],
    bert_scorer: BertScoreProxy | None,
) -> list[dict]:
    evaluated = []
    for row in rows:
        query = queries[row["query_id"]]
        reference_text = query["ground_truth_text"]
        generated_text = row["generated_text"]
        evaluated.append({
            **row,
            "category": query["category"],
            "reference_text": reference_text,
            "bleu": bleu4_score(generated_text, reference_text),
            "rouge_l": rouge_l_f1(generated_text, reference_text),
            "bertscore": bert_scorer.score(generated_text, reference_text) if bert_scorer else bertscore_fallback(generated_text, reference_text),
            "human_score": "",
        })
    return evaluated


def aggregate_rows(rows: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str, int], list[dict]] = defaultdict(list)
    for row in rows:
        grouped[(row["family"], row["system_label"], row["top_k"])].append(row)

    summary = []
    for (family, system_label, top_k), items in sorted(grouped.items(), key=lambda item: (item[0][2], item[0][0], item[0][1])):
        def mean(key: str) -> float:
            values = [float(x[key]) for x in items if x[key] != ""]
            return sum(values) / len(values) if values else 0.0

        summary.append({
            "family": family,
            "system_label": system_label,
            "top_k": top_k,
            "num_queries": len(items),
            "bleu": mean("bleu"),
            "rouge_l": mean("rouge_l"),
            "bertscore": mean("bertscore"),
            "human_score": "",
        })
    return summary


def build_dkg_rankings(summary_rows: list[dict]) -> list[dict]:
    dkg_rows = [row for row in summary_rows if row["family"] == "DKG"]
    ranked_by_metric = {
        "bleu_rank": sorted(dkg_rows, key=lambda row: row["bleu"], reverse=True),
        "rouge_l_rank": sorted(dkg_rows, key=lambda row: row["rouge_l"], reverse=True),
        "bertscore_rank": sorted(dkg_rows, key=lambda row: row["bertscore"], reverse=True),
    }

    rank_maps: dict[str, dict[tuple[str, int], int]] = {}
    for label, ranked_rows in ranked_by_metric.items():
        rank_maps[label] = {
            (row["system_label"], row["top_k"]): idx
            for idx, row in enumerate(ranked_rows, start=1)
        }

    rankings = []
    for row in dkg_rows:
        key = (row["system_label"], row["top_k"])
        rankings.append({
            **row,
            "bleu_rank": rank_maps["bleu_rank"][key],
            "rouge_l_rank": rank_maps["rouge_l_rank"][key],
            "bertscore_rank": rank_maps["bertscore_rank"][key],
            "avg_rank": (
                rank_maps["bleu_rank"][key]
                + rank_maps["rouge_l_rank"][key]
                + rank_maps["bertscore_rank"][key]
            ) / 3.0,
        })
    rankings.sort(key=lambda row: (row["avg_rank"], -row["bertscore"], -row["rouge_l"], -row["bleu"]))
    return rankings


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate DKG/TAARN/GPT generation results against human complaints")
    parser.add_argument("--sdkg-jsonl", nargs="*", default=[], help="One or more SDKG/DKG output jsonl files")
    parser.add_argument("--taarn-jsonl", nargs="*", default=[], help="One or more TAARN output jsonl files")
    parser.add_argument("--query-file", default=str(QUERY_FILE), help="0519 test sheet path")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Directory for evaluation CSV files")
    parser.add_argument("--output-stem", default="generation_eval_20lines", help="Output file stem")
    parser.add_argument("--skip-bertscore", action="store_true", help="Skip transformer-based BERTScore proxy")
    parser.add_argument("--bert-model", default=DEFAULT_BERT_MODEL, help="HF model name for BERTScore proxy")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    query_file = Path(args.query_file)
    output_dir = Path(args.output_dir)

    queries, _ = load_query_sheet(query_file)
    sdkg_paths = [Path(path) for path in args.sdkg_jsonl]
    taarn_paths = [Path(path) for path in args.taarn_jsonl]

    sdkg_rows = parse_sdkg_records(sdkg_paths)
    taarn_rows = parse_taarn_records(taarn_paths)
    top_k_values = sorted({row["top_k"] for row in sdkg_rows + taarn_rows if row["top_k"] > 0})
    if not top_k_values:
        top_k_values = [3]
    gpt_rows = build_gpt_rows(queries, top_k_values)

    bert_scorer = None if args.skip_bertscore else BertScoreProxy(args.bert_model)
    evaluated_rows = evaluate_rows(sdkg_rows + taarn_rows + gpt_rows, queries, bert_scorer)
    summary_rows = aggregate_rows(evaluated_rows)

    long_fields = [
        "family",
        "system_label",
        "query_id",
        "category",
        "top_k",
        "bleu",
        "rouge_l",
        "bertscore",
        "human_score",
        "source_file",
        "generated_text",
        "reference_text",
    ]
    summary_fields = [
        "family",
        "system_label",
        "top_k",
        "num_queries",
        "bleu",
        "rouge_l",
        "bertscore",
        "human_score",
    ]

    long_path = output_dir / f"{args.output_stem}_long.csv"
    summary_path = output_dir / f"{args.output_stem}_summary.csv"
    ranking_path = output_dir / f"{args.output_stem}_dkg_rankings.csv"
    write_csv(long_path, evaluated_rows, long_fields)
    write_csv(summary_path, summary_rows, summary_fields)
    write_csv(
        ranking_path,
        build_dkg_rankings(summary_rows),
        summary_fields + ["bleu_rank", "rouge_l_rank", "bertscore_rank", "avg_rank"],
    )

    print(f"saved_long_csv={long_path}")
    print(f"saved_summary_csv={summary_path}")
    print(f"saved_dkg_rankings_csv={ranking_path}")


if __name__ == "__main__":
    main()
