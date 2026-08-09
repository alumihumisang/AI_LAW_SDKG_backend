"""
SDKG Phase 1: 16-dim Boolean Encoding Matrix

Goal:
- Rebuild Phase 1 features strictly according to the 16-cell boolean matrix
- Reuse the already-corrected phase1_tensors_v4.jsonl as the source of texts and litigant hints
- Output a stable 16-dim boolean tensor for later retrieval / KG experiments

16 dimensions (fixed order):
1.  single_plaintiff
2.  multiple_plaintiffs
3.  single_defendant
4.  multiple_defendants
5.  negligence
6.  gross_negligence
7.  joint_liability
8.  prior_criminal
9.  head_neck
10. trunk
11. extremities
12. psych_other
13. medical_rehab
14. lost_income
15. non_pecuniary
16. care_other
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent


def existing_input_file(name: str) -> Path:
    local_path = BASE_DIR / name
    if not local_path.exists():
        raise FileNotFoundError(f"Required SDKG input file not found: {local_path}")
    return local_path


SRC_FILE = existing_input_file("phase1_tensors_v4.jsonl")
OUT_FILE = BASE_DIR / "phase1_boolean_matrix_v1.jsonl"
LAWYER_INPUT_FILE = BASE_DIR / "data" / "lawyer_inputs_6057.xlsx"


TENSOR_ORDER = [
    "single_plaintiff",
    "multiple_plaintiffs",
    "single_defendant",
    "multiple_defendants",
    "negligence",
    "gross_negligence",
    "joint_liability",
    "prior_criminal",
    "head_neck",
    "trunk",
    "extremities",
    "psych_other",
    "medical_rehab",
    "lost_income",
    "non_pecuniary",
    "care_other",
]


GROSS_NEG_PATTERNS = [
    r"酒[後醉]駕",
    r"吐氣酒精濃度",
    r"無照",
    r"駕照.*吊銷",
    r"闖紅燈",
    r"逆向",
    r"超速",
    r"兩段式左轉",
    r"未依.*閃光紅燈",
    r"明知.*仍",
]

JOINT_LIABILITY_PATTERNS = [
    r"連帶",
    r"共同不法侵害",
    r"共同侵權",
    r"民法第\s*18[78]\s*條",
    r"第\s*18[78]\s*條",
    r"法定代理人",
    r"僱用人",
    r"與其法定代理人連帶",
]

PRIOR_CRIMINAL_PATTERNS = [
    r"前科",
    r"刑事",
    r"緩起訴",
    r"有期徒刑",
    r"拘役",
    r"酒駕案件",
    r"吊銷.*駕照",
    r"遭註銷",
    r"刑案",
]

HEAD_NECK_PATTERNS = [
    r"頭",
    r"頭暈",
    r"頭痛",
    r"顱",
    r"腦",
    r"頸",
    r"頸椎",
    r"顏面",
    r"面部",
    r"臉部",
]

TRUNK_PATTERNS = [
    r"胸",
    r"腹",
    r"腰",
    r"下背",
    r"背部",
    r"脊椎",
    r"骨盆",
    r"軀幹",
]

EXTREMITY_PATTERNS = [
    r"肩",
    r"上肢",
    r"下肢",
    r"手",
    r"腕",
    r"手肘",
    r"手臂",
    r"前臂",
    r"腿",
    r"膝",
    r"膝蓋",
    r"踝",
    r"腳",
    r"足",
    r"四肢",
    r"鎖骨",
    r"橈骨",
    r"韌帶",
]

PSYCH_PATTERNS = [
    r"精神",
    r"憂鬱",
    r"失眠",
    r"安眠藥",
    r"抗憂鬱",
    r"精神科",
    r"輕生",
    r"創傷後",
    r"PTSD",
    r"心理",
]

MEDICAL_REHAB_PATTERNS = [
    r"醫療費",
    r"醫藥費",
    r"診療費",
    r"復健",
    r"門診",
    r"急診",
    r"藥品",
    r"醫療用品",
    r"護具",
    r"醫材",
    r"住院",
]

LOST_INCOME_PATTERNS = [
    r"薪資損失",
    r"工作損失",
    r"不能工作",
    r"無法工作",
    r"減少勞動能力",
    r"勞動能力",
    r"收入減少",
    r"請假",
    r"失去工作",
]

NON_PECUNIARY_PATTERNS = [
    r"慰撫金",
    r"精神慰撫金",
    r"非財產上",
]

CARE_OTHER_PATTERNS = [
    r"看護費",
    r"交通費",
    r"計程車",
    r"車輛",
    r"修理費",
    r"修復費",
    r"安全帽",
    r"營養品",
    r"照護",
    r"看護",
    r"護理",
]


def compile_any(patterns: List[str]) -> re.Pattern[str]:
    return re.compile("|".join(f"(?:{p})" for p in patterns))


GROSS_NEG_RE = compile_any(GROSS_NEG_PATTERNS)
JOINT_LIABILITY_RE = compile_any(JOINT_LIABILITY_PATTERNS)
PRIOR_CRIMINAL_RE = compile_any(PRIOR_CRIMINAL_PATTERNS)
HEAD_NECK_RE = compile_any(HEAD_NECK_PATTERNS)
TRUNK_RE = compile_any(TRUNK_PATTERNS)
EXTREMITY_RE = compile_any(EXTREMITY_PATTERNS)
PSYCH_RE = compile_any(PSYCH_PATTERNS)
MEDICAL_REHAB_RE = compile_any(MEDICAL_REHAB_PATTERNS)
LOST_INCOME_RE = compile_any(LOST_INCOME_PATTERNS)
NON_PECUNIARY_RE = compile_any(NON_PECUNIARY_PATTERNS)
CARE_OTHER_RE = compile_any(CARE_OTHER_PATTERNS)


def load_lawyer_inputs() -> Dict[str, str]:
    wb = load_workbook(LAWYER_INPUT_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx_case = header.index("case_id")
    idx_input = header.index("律師輸入")
    result: Dict[str, str] = {}
    for row in rows:
        if row[idx_case] is None:
            continue
        result[str(row[idx_case])] = "" if row[idx_input] is None else str(row[idx_input])
    return result


def iter_records(path: Path):
    with path.open(encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                yield json.loads(line)


def text_blob(*parts: str) -> str:
    return "\n".join(p for p in parts if p).strip()


def has_match(regex: re.Pattern[str], text: str) -> int:
    return 1 if text and regex.search(text) else 0


def extract_litigants(rec: dict) -> Dict[str, int]:
    lig = rec.get("litigant", {})
    return {
        "single_plaintiff": int(lig.get("L1", 0)),
        "multiple_plaintiffs": int(lig.get("L2", 0)),
        "single_defendant": int(lig.get("L3", 0)),
        "multiple_defendants": int(lig.get("L4", 0)),
    }


def extract_fact(rec: dict) -> Dict[str, int]:
    full = text_blob(rec.get("fact_text", ""), rec.get("laws_text", ""), rec.get("conclusion_text", ""))
    return {
        "negligence": 1,
        "gross_negligence": has_match(GROSS_NEG_RE, full),
        "joint_liability": has_match(JOINT_LIABILITY_RE, full),
        "prior_criminal": has_match(PRIOR_CRIMINAL_RE, full),
    }


def extract_injury(rec: dict) -> Dict[str, int]:
    injury_text = text_blob(
        rec.get("LI_injury", ""),
        rec.get("fact_text", ""),
        rec.get("compensation_text", ""),
        rec.get("conclusion_text", ""),
    )
    head = has_match(HEAD_NECK_RE, injury_text)
    trunk = has_match(TRUNK_RE, injury_text)
    ext = has_match(EXTREMITY_RE, injury_text)
    psych = has_match(PSYCH_RE, injury_text)
    # fallback bucket: unspecified / uncategorized injury also goes to other
    psych_other = 1 if psych or (head == 0 and trunk == 0 and ext == 0) else 0
    return {
        "head_neck": head,
        "trunk": trunk,
        "extremities": ext,
        "psych_other": psych_other,
    }


def extract_compensation(rec: dict) -> Dict[str, int]:
    comp_text = text_blob(
        rec.get("LI_compensation", ""),
        rec.get("compensation_text", ""),
        rec.get("conclusion_text", ""),
    )
    med = has_match(MEDICAL_REHAB_RE, comp_text)
    income = has_match(LOST_INCOME_RE, comp_text)
    nonpec = has_match(NON_PECUNIARY_RE, comp_text)
    care = has_match(CARE_OTHER_RE, comp_text)
    # fallback bucket: unspecified / uncategorized compensation also goes to other
    care_other = 1 if care or (med == 0 and income == 0 and nonpec == 0) else 0
    return {
        "medical_rehab": med,
        "lost_income": income,
        "non_pecuniary": nonpec,
        "care_other": care_other,
    }


def build_review_flags(features: Dict[str, int], rec: dict) -> List[str]:
    flags: List[str] = []
    if features["single_plaintiff"] and features["multiple_plaintiffs"]:
        flags.append("plaintiff_count_conflict")
    if features["single_defendant"] and features["multiple_defendants"]:
        flags.append("defendant_count_conflict")
    if not (features["single_plaintiff"] or features["multiple_plaintiffs"]):
        flags.append("missing_plaintiff_count")
    if not (features["single_defendant"] or features["multiple_defendants"]):
        flags.append("missing_defendant_count")
    if features["gross_negligence"] and not features["negligence"]:
        flags.append("gross_without_negligence")
    if all(features[k] == 0 for k in ["medical_rehab", "lost_income", "non_pecuniary", "care_other"]):
        flags.append("missing_compensation_signal")
    if all(features[k] == 0 for k in ["head_neck", "trunk", "extremities", "psych_other"]):
        flags.append("missing_injury_signal")
    if rec.get("compensation_text") and len(rec.get("compensation_text", "")) > 30 and features["care_other"] == 1 and features["medical_rehab"] == 0 and features["lost_income"] == 0 and features["non_pecuniary"] == 0:
        flags.append("only_other_compensation")
    return flags


def tensor_from_features(features: Dict[str, int]) -> List[int]:
    return [int(features[name]) for name in TENSOR_ORDER]


def main() -> None:
    lawyer_inputs = load_lawyer_inputs()
    records = list(iter_records(SRC_FILE))
    print(f"Loaded source records: {len(records)}")

    with OUT_FILE.open("w", encoding="utf-8") as out:
        for rec in records:
            cid = str(rec["case_id"])
            features = {}
            features.update(extract_litigants(rec))
            features.update(extract_fact(rec))
            injury_features = extract_injury(rec)
            compensation_features = extract_compensation(rec)

            # Legal-practice alignment rule:
            # if a case claims non-pecuniary damages, Psych & Other must be on.
            if compensation_features["non_pecuniary"] == 1:
                injury_features["psych_other"] = 1

            features.update(injury_features)
            features.update(compensation_features)

            tensor = tensor_from_features(features)
            review_flags = build_review_flags(features, rec)

            output = {
                "case_id": cid,
                "lawyer_input": lawyer_inputs.get(cid, ""),
                "boolean_matrix": {
                    "litigants": {k: features[k] for k in TENSOR_ORDER[0:4]},
                    "fact": {k: features[k] for k in TENSOR_ORDER[4:8]},
                    "injury": {k: features[k] for k in TENSOR_ORDER[8:12]},
                    "compensation": {k: features[k] for k in TENSOR_ORDER[12:16]},
                },
                "tensor_order": TENSOR_ORDER,
                "tensor": tensor,
                "fact_text": rec.get("fact_text", ""),
                "laws_text": rec.get("laws_text", ""),
                "compensation_text": rec.get("compensation_text", ""),
                "conclusion_text": rec.get("conclusion_text", ""),
                "LI_fact": rec.get("LI_fact", ""),
                "LI_injury": rec.get("LI_injury", ""),
                "LI_compensation": rec.get("LI_compensation", ""),
                "review_flags": review_flags,
                "legacy_hints": {
                    "plaintiff_names": rec.get("plaintiff_names", []),
                    "defendant_names": rec.get("defendant_names", []),
                    "case_type": rec.get("case_type", []),
                    "has_vicarious": rec.get("has_vicarious", False),
                    "has_criminal_case": rec.get("has_criminal_case", False),
                },
            }
            out.write(json.dumps(output, ensure_ascii=False) + "\n")

    print(f"Wrote boolean matrix records to: {OUT_FILE}")


if __name__ == "__main__":
    main()
