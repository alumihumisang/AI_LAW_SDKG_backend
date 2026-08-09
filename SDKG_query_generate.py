from __future__ import annotations

"""
SDKG query generation pipeline

Architecture:
1. Query loading / corpus loading
2. Query feature extraction and SDKG retrieval
3. Legal-generation support rules and prompts
4. Section-wise lawsuit generation
5. Batch export for experiment evaluation

The legal extraction rules and strict prompt templates are intentionally
kept in `sdkg_generation_legal.py` so this file remains the orchestration
layer for the 50-query experiment pipeline.
"""

import argparse
import json
import csv
import re
import sys
from dataclasses import dataclass
from itertools import permutations
from pathlib import Path
from typing import Iterable

import numpy as np
import requests
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from sdkg_distance import (
    legal_feature_arrays,
    legal_feature_vector,
    shared_activated_feature_mask,
    weighted_legal_feature_distance,
)
from sdkg_generation_legal import (
    build_strict_facts_prompt,
    determine_applicable_laws_structured,
    extract_parties_structured,
)
from sdkg_generation_sections import (
    build_conclusion_section,
    build_damages_prompt,
    build_generation_support_context,
    build_structured_damage_section,
    clean_damage_section,
    extract_damage_constraints,
)


def existing_input_file(*parts: str) -> Path:
    local_path = BASE_DIR.joinpath(*parts)
    if not local_path.exists():
        raise FileNotFoundError(f"Required SDKG input file not found: {local_path}")
    return local_path


QUERY_FILE = existing_input_file("data", "queries_50_0519.xlsx")
COMPLAINT_FILE = existing_input_file("data", "complaints_6057.xlsx")
CASE_FILE = existing_input_file("phase1_boolean_severity_v1.jsonl")
LINKS_FILE = BASE_DIR / "experiment_outputs" / "experiment_links.csv"
SEVERITY_TREE_LINKS_FILE = existing_input_file(
    "experiment_outputs",
    "severity_trees",
    "severity_tree_parent_links_lambda_0p50.csv",
)
OUTPUT_DIR = BASE_DIR / "generation_outputs"
LLM_URL = "http://localhost:11434/api/generate"
DEFAULT_MODEL = "gemma3:27b"
ACTIVE_LAMBDA_LEGAL = 0.5

WEIGHT_PERMUTATIONS = sorted(set(permutations((0.5, 0.3, 0.2), 3)), reverse=True)
DISTANCE_THRESHOLDS = [0.100, 0.250, 0.500]

GROSS_NEG_PATTERNS = [
    r"酒[後醉]駕",
    r"吐氣酒精濃度",
    r"無照",
    r"駕照.*吊銷",
    r"闖紅燈",
    r"逆向",
    r"超速",
    r"兩段式左轉",
    r"未依.*閃光紅燈",
    r"明知.*仍",
]
JOINT_LIABILITY_PATTERNS = [
    r"連帶",
    r"共同不法侵害",
    r"共同侵權",
    r"民法第\s*18[78]\s*條",
    r"第\s*18[78]\s*條",
    r"法定代理人",
    r"僱用人",
    r"與其法定代理人連帶",
]
PRIOR_CRIMINAL_PATTERNS = [
    r"前科",
    r"刑事",
    r"緩起訴",
    r"有期徒刑",
    r"拘役",
    r"酒駕案件",
    r"吊銷.*駕照",
    r"遭註銷",
    r"刑案",
]
HEAD_NECK_PATTERNS = [r"頭", r"頭暈", r"頭痛", r"顱", r"腦", r"頸", r"頸椎", r"顏面", r"面部", r"臉部"]
TRUNK_PATTERNS = [r"胸", r"腹", r"腰", r"下背", r"背部", r"脊椎", r"骨盆", r"軀幹"]
EXTREMITY_PATTERNS = [
    r"肩", r"上肢", r"下肢", r"手", r"腕", r"手肘", r"手臂", r"前臂", r"腿", r"膝", r"膝蓋", r"踝", r"腳", r"足",
    r"四肢", r"鎖骨", r"橈骨", r"韌帶",
]
PSYCH_PATTERNS = [r"精神", r"憂鬱", r"失眠", r"安眠藥", r"抗憂鬱", r"精神科", r"輕生", r"創傷後", r"PTSD", r"心理"]
MEDICAL_REHAB_PATTERNS = [r"醫療費", r"醫藥費", r"診療費", r"復健", r"門診", r"急診", r"藥品", r"醫療用品", r"護具", r"醫材", r"住院"]
LOST_INCOME_PATTERNS = [r"薪資損失", r"工作損失", r"不能工作", r"無法工作", r"減少勞動能力", r"勞動能力", r"收入減少", r"請假", r"失去工作"]
NON_PECUNIARY_PATTERNS = [r"慰撫金", r"精神慰撫金", r"非財產上"]
CARE_OTHER_PATTERNS = [r"看護費", r"交通費", r"計程車", r"車輛", r"修理費", r"修復費", r"安全帽", r"營養品", r"照護", r"看護", r"護理"]

A_FACT_PATTERNS = [r"酒[後醉]駕", r"吐氣酒精濃度", r"無照", r"駕照.*吊銷", r"駕照.*註銷"]
B_FACT_PATTERNS = [r"闖紅燈", r"逆向", r"超速", r"肇事逃逸", r"兩段式左轉", r"未依.*燈號"]
C_FACT_PATTERNS = [r"未保持安全距離", r"追撞", r"未讓直行車", r"轉彎未讓", r"未禮讓", r"連帶", r"共同侵權", r"法定代理人", r"僱用人"]
E_FACT_PATTERNS = [r"閃避不及", r"猝不及防"]
A_INJURY_PATTERNS = [r"植物人", r"截肢", r"全身癱瘓", r"無法脫離呼吸器"]
B_INJURY_PATTERNS = [r"顱內出血", r"硬腦膜", r"脊椎骨折", r"脊髓", r"失明", r"癱瘓", r"需專人.*看護", r"長期看護", r"顱骨缺損"]
C_INJURY_PATTERNS = [r"骨折", r"鋼釘", r"住院", r"手術", r"開刀", r"粉碎性", r"韌帶撕裂"]
D_INJURY_PATTERNS = [r"挫傷", r"腦震盪", r"復健", r"拉傷", r"扭傷", r"骨裂"]
E_INJURY_PATTERNS = [r"擦傷", r"破皮", r"瘀青", r"瘀傷", r"擦挫傷"]
INSURANCE_DED_PATTERNS = [r"強制險", r"已獲賠償", r"已領.*保險", r"保險公司", r"扣除"]
TOTAL_AMOUNT_PATTERNS = [r"(?:合計|總計|共計|請求(?:被告)?(?:給付|賠償)?|賠償(?:金額)?(?:總計)?)\s*([0-9,]+)\s*元"]

FACT_SCORE = {"A": 1.0, "B": 0.8, "C": 0.6, "D": 0.4, "E": 0.2, "N": 0.0}
INJURY_SCORE = {"A": 1.0, "B": 0.8, "C": 0.6, "D": 0.4, "E": 0.2, "N": 0.0}
COMP_SCORE = {"A": 1.0, "B": 0.75, "C": 0.5, "D": 0.25, "N": 0.0}
INSURANCE_DEDUCTION_PENALTY = 0.2

TAU_CODE_BY_VALUE = {
    0.100: "L",
    0.250: "M",
    0.500: "H",
}

WEIGHT_CODE_BY_VALUES = {
    (0.5, 0.3, 0.2): "FI",
    (0.5, 0.2, 0.3): "FC",
    (0.3, 0.5, 0.2): "IF",
    (0.3, 0.2, 0.5): "CF",
    (0.2, 0.5, 0.3): "IC",
    (0.2, 0.3, 0.5): "CI",
}


@dataclass
class ExperimentTreeContext:
    exp_id: str
    heavy_parent_map: dict[str, str]
    heavy_child_map: dict[str, list[str]]
    light_parent_map: dict[str, str]
    light_child_map: dict[str, list[str]]


EXPERIMENT_TREE_CACHE: dict[str, ExperimentTreeContext] = {}


def compile_any(patterns: list[str]) -> re.Pattern[str]:
    return re.compile("|".join(f"(?:{p})" for p in patterns))


GROSS_NEG_RE = compile_any(GROSS_NEG_PATTERNS)
JOINT_LIABILITY_RE = compile_any(JOINT_LIABILITY_PATTERNS)
PRIOR_CRIMINAL_RE = compile_any(PRIOR_CRIMINAL_PATTERNS)
HEAD_NECK_RE = compile_any(HEAD_NECK_PATTERNS)
TRUNK_RE = compile_any(TRUNK_PATTERNS)
EXTREMITY_RE = compile_any(EXTREMITY_PATTERNS)
PSYCH_RE = compile_any(PSYCH_PATTERNS)
MEDICAL_REHAB_RE = compile_any(MEDICAL_REHAB_PATTERNS)
LOST_INCOME_RE = compile_any(LOST_INCOME_PATTERNS)
NON_PECUNIARY_RE = compile_any(NON_PECUNIARY_PATTERNS)
CARE_OTHER_RE = compile_any(CARE_OTHER_PATTERNS)

A_FACT_RE = compile_any(A_FACT_PATTERNS)
B_FACT_RE = compile_any(B_FACT_PATTERNS)
C_FACT_RE = compile_any(C_FACT_PATTERNS)
E_FACT_RE = compile_any(E_FACT_PATTERNS)
A_INJURY_RE = compile_any(A_INJURY_PATTERNS)
B_INJURY_RE = compile_any(B_INJURY_PATTERNS)
C_INJURY_RE = compile_any(C_INJURY_PATTERNS)
D_INJURY_RE = compile_any(D_INJURY_PATTERNS)
E_INJURY_RE = compile_any(E_INJURY_PATTERNS)
INSURANCE_DED_RE = compile_any(INSURANCE_DED_PATTERNS)


def iter_records(path: Path) -> Iterable[dict]:
    with path.open(encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                yield json.loads(line)


def load_case_corpus(
    case_limit: int | None = None,
) -> tuple[
    list[dict],
    np.ndarray,
    dict[str, np.ndarray],
    dict[str, list[str]],
    np.ndarray,
    np.ndarray,
    np.ndarray,
    np.ndarray,
]:
    complaint_texts = load_full_complaints()
    rows = []
    litigant_values = []
    fact_values = []
    injury_values = []
    comp_values = []
    case_sort = []
    for idx, rec in enumerate(iter_records(CASE_FILE)):
        if case_limit is not None and idx >= case_limit:
            break
        rec = dict(rec)
        rec["complaint_text"] = complaint_texts.get(str(rec.get("case_id")), build_case_full_text(rec))
        rows.append(rec)
        litigants = rec["boolean_matrix"]["litigants"]
        litigant_values.append([
            float(litigants.get("single_plaintiff", 0)),
            float(litigants.get("multiple_plaintiffs", 0)),
            float(litigants.get("single_defendant", 0)),
            float(litigants.get("multiple_defendants", 0)),
        ])
        scores = rec["severity_scores"]
        fact_values.append(float(scores["Fact"]))
        injury_values.append(float(scores["Injury"]))
        comp_values.append(float(scores["Compensation"]))
        cid = str(rec["case_id"])
        case_sort.append(int(cid) if cid.isdigit() else 10**12 + len(case_sort))
    legal_features, legal_keys = legal_feature_arrays(rows)

    return (
        rows,
        np.array(litigant_values, dtype=np.float32),
        legal_features,
        legal_keys,
        np.array(fact_values, dtype=np.float32),
        np.array(injury_values, dtype=np.float32),
        np.array(comp_values, dtype=np.float32),
        np.array(case_sort, dtype=np.int64),
    )


def load_full_complaints() -> dict[str, str]:
    if not COMPLAINT_FILE.exists():
        return {}
    wb = load_workbook(COMPLAINT_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: pos for pos, name in enumerate(header)}
    if "case_id" not in idx or "起訴書" not in idx:
        return {}
    result: dict[str, str] = {}
    for row in rows:
        case_id = row[idx["case_id"]]
        complaint = row[idx["起訴書"]]
        if case_id in (None, "") or complaint in (None, ""):
            continue
        result[str(case_id)] = str(complaint)
    return result


def build_case_full_text(rec: dict) -> str:
    complaint_text = str(rec.get("complaint_text", "") or "").strip()
    if complaint_text:
        return complaint_text
    parts = [
        rec.get("fact_text", ""),
        rec.get("compensation_text", ""),
        rec.get("conclusion_text", ""),
    ]
    return "\n\n".join(str(part).strip() for part in parts if str(part).strip())


def load_parent_map(exp_id: str) -> dict[str, str]:
    parent_map: dict[str, str] = {}
    if not LINKS_FILE.exists():
        return parent_map
    with LINKS_FILE.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["exp_id"] == exp_id and row["link_type"] == "parent":
                parent_map[str(row["target_case_id"])] = str(row["source_case_id"])
    return parent_map


def load_severity_tree_parent_maps(exp_id: str) -> tuple[dict[str, str], dict[str, str]]:
    lh_parent_map: dict[str, str] = {}
    hl_parent_map: dict[str, str] = {}
    if not SEVERITY_TREE_LINKS_FILE.exists():
        return lh_parent_map, hl_parent_map
    with SEVERITY_TREE_LINKS_FILE.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["exp_id"] != exp_id:
                continue
            parent_map = lh_parent_map if row["direction"] == "LH" else hl_parent_map
            parent_map[str(row["child_case_id"])] = str(row["parent_case_id"])
    return lh_parent_map, hl_parent_map


def filter_parent_map(parent_map: dict[str, str], case_ids: set[str]) -> dict[str, str]:
    return {
        child_id: parent_id
        for child_id, parent_id in parent_map.items()
        if child_id in case_ids and parent_id in case_ids
    }


def build_child_map(parent_map: dict[str, str]) -> dict[str, list[str]]:
    child_map: dict[str, list[str]] = {}
    for child_id, parent_id in parent_map.items():
        child_map.setdefault(parent_id, []).append(child_id)
    return child_map


def compute_reverse_parent_map(
    exp: dict,
    case_ids: list[str],
    case_sort: np.ndarray,
    litigant_values: np.ndarray,
    legal_features: dict[str, np.ndarray],
    fact_values: np.ndarray,
    injury_values: np.ndarray,
    comp_values: np.ndarray,
) -> dict[str, str]:
    fact_w = np.float32(exp["fact_w"])
    injury_w = np.float32(exp["injury_w"])
    comp_w = np.float32(exp["comp_w"])
    tau = np.float32(exp["distance_threshold"])

    scores = fact_w * fact_values + injury_w * injury_values + comp_w * comp_values
    df = weighted_legal_feature_distance(
        legal_features,
        legal_features,
        fact_w,
        injury_w,
        comp_w,
    )
    shared_feature = shared_activated_feature_mask(legal_features, legal_features)
    ds = (
        fact_w * np.abs(fact_values[:, None] - fact_values[None, :])
        + injury_w * np.abs(injury_values[:, None] - injury_values[None, :])
        + comp_w * np.abs(comp_values[:, None] - comp_values[None, :])
    )
    lam = np.float32(ACTIVE_LAMBDA_LEGAL)
    distance_matrix = (lam * df + (np.float32(1.0) - lam) * ds).astype(np.float32)
    np.fill_diagonal(distance_matrix, np.float32(np.inf))

    reverse_parent_map: dict[str, str] = {}
    root_idx = int(np.argmin(scores))
    for i, cid in enumerate(case_ids):
        if i == root_idx:
            continue
        drow = distance_matrix[i]
        candidate_idx = np.where((scores < scores[i]) & (drow <= tau) & shared_feature[i])[0]
        if candidate_idx.size == 0:
            continue
        ordered_idx = candidate_idx[
            np.lexsort(
                (
                    case_sort[candidate_idx],
                    scores[candidate_idx],
                    drow[candidate_idx],
                )
            )
        ]
        parent_idx = int(ordered_idx[0])
        reverse_parent_map[cid] = case_ids[parent_idx]

    return reverse_parent_map


def get_experiment_tree_context(
    exp: dict,
    case_ids: list[str],
    case_sort: np.ndarray,
    litigant_values: np.ndarray,
    legal_features: dict[str, np.ndarray],
    fact_values: np.ndarray,
    injury_values: np.ndarray,
    comp_values: np.ndarray,
) -> ExperimentTreeContext:
    exp_id = exp["exp_id"]
    cache_key = f"{exp_id}|{len(case_ids)}|{case_ids[0] if case_ids else ''}|{case_ids[-1] if case_ids else ''}|{SEVERITY_TREE_LINKS_FILE}"
    cached = EXPERIMENT_TREE_CACHE.get(cache_key)
    if cached is not None:
        return cached

    light_parent_map, heavy_parent_map = load_severity_tree_parent_maps(exp_id)
    case_id_set = set(case_ids)
    light_parent_map = filter_parent_map(light_parent_map, case_id_set)
    heavy_parent_map = filter_parent_map(heavy_parent_map, case_id_set)
    if not light_parent_map or not heavy_parent_map:
        heavy_parent_map = load_parent_map(exp_id)
        heavy_parent_map = filter_parent_map(heavy_parent_map, case_id_set)
        light_parent_map = compute_reverse_parent_map(
            exp,
            case_ids,
            case_sort,
            litigant_values,
            legal_features,
            fact_values,
            injury_values,
            comp_values,
        )
    context = ExperimentTreeContext(
        exp_id=exp_id,
        heavy_parent_map=heavy_parent_map,
        heavy_child_map=build_child_map(heavy_parent_map),
        light_parent_map=light_parent_map,
        light_child_map=build_child_map(light_parent_map),
    )
    EXPERIMENT_TREE_CACHE[cache_key] = context
    return context


def load_queries(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: header.index(name) for name in header}
    items = []
    current_category = ""
    for row in rows:
        query_id = row[idx["query_id"]]
        query_text = row[idx["模擬輸入內容"]]
        if not query_id or not query_text:
            continue
        raw_category = row[idx["類別"]]
        if raw_category not in (None, ""):
            current_category = str(raw_category)
        items.append({
            "category": current_category,
            "query_id": int(query_id),
            "query_text": str(query_text),
            "reference_text": str(row[idx["Ground Truth-人工起訴書"]] or ""),
            "ground_truth_text": str(row[idx["Ground Truth-人工起訴書"]] or ""),
            "gpt_baseline_text": str(row[idx["gpt-4o-mini"]] or ""),
        })
    return items


def extract_sections(text: str) -> dict[str, str]:
    result = {"accident_facts": "", "injuries": "", "compensation_facts": ""}
    fact_match = re.search(r"一[、．.\s]*事故發生緣由[:：]?\s*(.*?)(?=二[、．.]|$)", text, re.S)
    injury_match = re.search(r"二[、．.\s]*(?:原告)?受傷情形[:：]?\s*(.*?)(?=三[、．.]|$)", text, re.S)
    comp_match = re.search(r"三[、．.\s]*請求賠償的事實根據[:：]?\s*(.*?)$", text, re.S)
    if fact_match:
        result["accident_facts"] = fact_match.group(1).strip()
    if injury_match:
        result["injuries"] = injury_match.group(1).strip()
    if comp_match:
        result["compensation_facts"] = comp_match.group(1).strip()
    return result


def extract_parties(query_text: str, model: str = DEFAULT_MODEL) -> dict:
    return extract_parties_structured(query_text, LLM_URL, model)


def has_match(regex: re.Pattern[str], text: str) -> int:
    return 1 if text and regex.search(text) else 0


def build_query_boolean_matrix(query_text: str, category: str) -> dict:
    sections = extract_sections(query_text)
    full_text = "\n".join(v for v in [query_text, sections["accident_facts"], sections["injuries"], sections["compensation_facts"]] if v)
    litigants = {
        "single_plaintiff": 1,
        "multiple_plaintiffs": 0,
        "single_defendant": 1,
        "multiple_defendants": 0,
    }
    if "原被告皆數名" in category:
        litigants = {"single_plaintiff": 0, "multiple_plaintiffs": 1, "single_defendant": 0, "multiple_defendants": 1}
    elif "數名原告" in category:
        litigants = {"single_plaintiff": 0, "multiple_plaintiffs": 1, "single_defendant": 1, "multiple_defendants": 0}
    elif "數名被告" in category:
        litigants = {"single_plaintiff": 1, "multiple_plaintiffs": 0, "single_defendant": 0, "multiple_defendants": 1}

    return {
        "litigants": litigants,
        "fact": {
            "negligence": 1,
            "gross_negligence": has_match(GROSS_NEG_RE, full_text),
            "joint_liability": 1 if ("§187" in category or "§188" in category or "§190" in category or JOINT_LIABILITY_RE.search(full_text)) else 0,
            "prior_criminal": has_match(PRIOR_CRIMINAL_RE, full_text),
        },
        "injury": {
            "head_neck": has_match(HEAD_NECK_RE, sections["injuries"] or full_text),
            "trunk": has_match(TRUNK_RE, sections["injuries"] or full_text),
            "extremities": has_match(EXTREMITY_RE, sections["injuries"] or full_text),
            "psych_other": has_match(PSYCH_RE, full_text),
        },
        "compensation": {
            "medical_rehab": has_match(MEDICAL_REHAB_RE, sections["compensation_facts"] or full_text),
            "lost_income": has_match(LOST_INCOME_RE, sections["compensation_facts"] or full_text),
            "non_pecuniary": has_match(NON_PECUNIARY_RE, sections["compensation_facts"] or full_text),
            "care_other": has_match(CARE_OTHER_RE, sections["compensation_facts"] or full_text),
        },
    }


def extract_total_amount(text: str) -> int:
    amounts: list[int] = []
    for pattern in TOTAL_AMOUNT_PATTERNS:
        for match in re.finditer(pattern, text):
            raw = match.group(1).replace(",", "")
            try:
                amounts.append(int(raw))
            except ValueError:
                continue
    if amounts:
        return max(amounts)
    fallback = []
    for match in re.finditer(r"([0-9][0-9,]{3,})\s*元", text):
        raw = match.group(1).replace(",", "")
        try:
            fallback.append(int(raw))
        except ValueError:
            continue
    return max(fallback) if fallback else 0


def amount_to_comp_level(amount: int) -> str:
    if amount > 2_500_000:
        return "A"
    if amount > 1_000_000:
        return "B"
    if amount > 300_000:
        return "C"
    if amount > 0:
        return "D"
    return "N"


def classify_query_levels(query_text: str, boolean_matrix: dict) -> tuple[dict, dict]:
    sections = extract_sections(query_text)
    fact_text = sections["accident_facts"] or query_text
    injury_text = sections["injuries"] or query_text
    comp_text = "\n".join(v for v in [sections["compensation_facts"], query_text] if v)
    fact_flags = boolean_matrix["fact"]
    injury_flags = boolean_matrix["injury"]
    comp_flags = boolean_matrix["compensation"]

    if fact_flags.get("gross_negligence") and A_FACT_RE.search(fact_text):
        fact_level = "A"
    elif fact_flags.get("gross_negligence"):
        fact_level = "B"
    elif fact_flags.get("joint_liability") or C_FACT_RE.search(fact_text):
        fact_level = "C"
    elif fact_flags.get("negligence"):
        fact_level = "D"
    elif E_FACT_RE.search(fact_text):
        fact_level = "E"
    else:
        fact_level = "N"

    if A_INJURY_RE.search(injury_text):
        injury_level = "A"
    elif B_INJURY_RE.search(injury_text):
        injury_level = "B"
    elif C_INJURY_RE.search(injury_text):
        injury_level = "C"
    elif D_INJURY_RE.search(injury_text):
        injury_level = "D"
    elif E_INJURY_RE.search(injury_text):
        injury_level = "E"
    elif injury_flags.get("head_neck") or injury_flags.get("trunk") or injury_flags.get("extremities"):
        injury_level = "D"
    elif injury_flags.get("psych_other"):
        injury_level = "E"
    else:
        injury_level = "N"

    total_amount = extract_total_amount(comp_text)
    comp_level = amount_to_comp_level(total_amount)
    if comp_level == "N" and any(comp_flags.get(k, 0) for k in ["medical_rehab", "lost_income", "non_pecuniary", "care_other"]):
        comp_level = "D"
    has_deduction = bool(INSURANCE_DED_RE.search(comp_text))
    base_comp = COMP_SCORE[comp_level]
    final_comp = base_comp - INSURANCE_DEDUCTION_PENALTY if has_deduction else base_comp

    return (
        {"Fact": fact_level, "Injury": injury_level, "Compensation": comp_level},
        {"Fact": FACT_SCORE[fact_level], "Injury": INJURY_SCORE[injury_level], "Compensation": final_comp},
    )


def build_experiments() -> list[dict]:
    experiments = []
    exp_num = 1
    for threshold in DISTANCE_THRESHOLDS:
        tau_code = TAU_CODE_BY_VALUE[float(threshold)]
        for fact_w, injury_w, comp_w in WEIGHT_PERMUTATIONS:
            weight_code = WEIGHT_CODE_BY_VALUES[(float(fact_w), float(injury_w), float(comp_w))]
            experiments.append({
                "exp_id": f"E{exp_num:02d}",
                "fact_w": float(fact_w),
                "injury_w": float(injury_w),
                "comp_w": float(comp_w),
                "distance_threshold": float(threshold),
                "weight_code": weight_code,
                "tau_code": tau_code,
                "short_name": f"{weight_code}-{tau_code}",
            })
            exp_num += 1
    return experiments


def retrieve_similar_cases(query_row: dict, exp: dict, corpus_rows: list[dict], litigant_values: np.ndarray, legal_features: dict[str, np.ndarray], legal_keys: dict[str, list[str]], fact_values: np.ndarray, injury_values: np.ndarray, comp_values: np.ndarray, case_sort: np.ndarray, top_k: int) -> list[dict]:
    boolean_matrix = build_query_boolean_matrix(query_row["query_text"], query_row["category"])
    _, query_scores = classify_query_levels(query_row["query_text"], boolean_matrix)

    fact_w = np.float32(exp["fact_w"])
    injury_w = np.float32(exp["injury_w"])
    comp_w = np.float32(exp["comp_w"])
    tau = np.float32(exp["distance_threshold"])
    q_fact = np.float32(query_scores["Fact"])
    q_injury = np.float32(query_scores["Injury"])
    q_comp = np.float32(query_scores["Compensation"])
    q_score = float(fact_w * q_fact + injury_w * q_injury + comp_w * q_comp)

    query_legal_features = legal_feature_vector(boolean_matrix, legal_keys)
    df = weighted_legal_feature_distance(
        query_legal_features,
        legal_features,
        fact_w,
        injury_w,
        comp_w,
    )[0]
    ds = (
        fact_w * np.abs(fact_values - q_fact)
        + injury_w * np.abs(injury_values - q_injury)
        + comp_w * np.abs(comp_values - q_comp)
    )
    lam = np.float32(ACTIVE_LAMBDA_LEGAL)
    distances = (lam * df + (np.float32(1.0) - lam) * ds).astype(np.float32)
    shared_feature = shared_activated_feature_mask(query_legal_features, legal_features)[0]

    candidate_idx = np.where((distances <= tau) & shared_feature)[0]
    if candidate_idx.size < top_k:
        fallback_idx = np.where(shared_feature)[0]
        candidate_idx = fallback_idx if fallback_idx.size else np.arange(len(corpus_rows))
    ordered_idx = candidate_idx[np.lexsort((case_sort[candidate_idx], -(
        fact_w * fact_values[candidate_idx] + injury_w * injury_values[candidate_idx] + comp_w * comp_values[candidate_idx]
    ), distances[candidate_idx]))][:top_k]

    result = []
    for rank, idx in enumerate(ordered_idx, start=1):
        rec = corpus_rows[int(idx)]
        case_score = float(fact_w * fact_values[idx] + injury_w * injury_values[idx] + comp_w * comp_values[idx])
        result.append({
            "rank": rank,
            "case_id": str(rec["case_id"]),
            "distance": float(distances[idx]),
            "case_score": case_score,
            "query_score": q_score,
            "fact_value": float(rec["severity_scores"]["Fact"]),
            "injury_value": float(rec["severity_scores"]["Injury"]),
            "comp_value": float(rec["severity_scores"]["Compensation"]),
            "fact_text": rec.get("LI_fact") or rec.get("fact_text") or "",
            "injury_text": rec.get("LI_injury") or "",
            "comp_text": rec.get("LI_compensation") or rec.get("compensation_text") or "",
            "conclusion_text": rec.get("conclusion_text") or "",
            "case_full_text": build_case_full_text(rec),
        })
    return result


def retrieve_dual_tree_cases(
    query_row: dict,
    exp: dict,
    corpus_rows: list[dict],
    litigant_values: np.ndarray,
    legal_features: dict[str, np.ndarray],
    legal_keys: dict[str, list[str]],
    fact_values: np.ndarray,
    injury_values: np.ndarray,
    comp_values: np.ndarray,
    case_sort: np.ndarray,
    top_k: int,
    tree_context: ExperimentTreeContext,
    case_idx_by_id: dict[str, int],
) -> tuple[list[dict], dict]:
    boolean_matrix = build_query_boolean_matrix(query_row["query_text"], query_row["category"])
    _, query_scores = classify_query_levels(query_row["query_text"], boolean_matrix)

    fact_w = np.float32(exp["fact_w"])
    injury_w = np.float32(exp["injury_w"])
    comp_w = np.float32(exp["comp_w"])
    q_fact = np.float32(query_scores["Fact"])
    q_injury = np.float32(query_scores["Injury"])
    q_comp = np.float32(query_scores["Compensation"])
    q_score = float(fact_w * q_fact + injury_w * q_injury + comp_w * q_comp)

    query_legal_features = legal_feature_vector(boolean_matrix, legal_keys)
    df = weighted_legal_feature_distance(
        query_legal_features,
        legal_features,
        fact_w,
        injury_w,
        comp_w,
    )[0]
    ds = (
        fact_w * np.abs(fact_values - q_fact)
        + injury_w * np.abs(injury_values - q_injury)
        + comp_w * np.abs(comp_values - q_comp)
    )
    lam = np.float32(ACTIVE_LAMBDA_LEGAL)
    distances = (lam * df + (np.float32(1.0) - lam) * ds).astype(np.float32)
    shared_feature = shared_activated_feature_mask(query_legal_features, legal_features)[0]
    case_scores = (fact_w * fact_values + injury_w * injury_values + comp_w * comp_values).astype(np.float32)
    anchor_candidates = np.where(shared_feature)[0]
    if anchor_candidates.size == 0:
        anchor_candidates = np.arange(len(corpus_rows))
    ordered_global_idx = anchor_candidates[
        np.lexsort((
            case_sort[anchor_candidates],
            case_scores[anchor_candidates] * 0 - case_scores[anchor_candidates],
            distances[anchor_candidates],
        ))
    ]
    anchor_idx = int(ordered_global_idx[0])
    anchor_case_id = str(corpus_rows[anchor_idx]["case_id"])

    retrieval_target = max(top_k, 8)
    directional_target = max(retrieval_target - 1, 0)
    lighter_target = directional_target // 2
    heavier_target = directional_target - lighter_target

    lighter_case_ids = collect_descendant_cases(
        anchor_case_id,
        tree_context.heavy_child_map,
        lighter_target,
        distances,
        case_idx_by_id,
    )
    heavier_case_ids = collect_descendant_cases(
        anchor_case_id,
        tree_context.light_child_map,
        heavier_target,
        distances,
        case_idx_by_id,
    )

    selected_case_ids: list[tuple[str, str]] = [(anchor_case_id, "anchor")]
    seen = {anchor_case_id}
    for cid in lighter_case_ids:
        if cid not in seen:
            selected_case_ids.append((cid, "lighter_tree"))
            seen.add(cid)
    for cid in heavier_case_ids:
        if cid not in seen:
            selected_case_ids.append((cid, "heavier_tree"))
            seen.add(cid)

    for idx in ordered_global_idx[1:]:
        cid = str(corpus_rows[int(idx)]["case_id"])
        if cid in seen:
            continue
        selected_case_ids.append((cid, "fallback"))
        seen.add(cid)
        if len(selected_case_ids) >= retrieval_target:
            break

    ranked_results = []
    for rank, (cid, source_side) in enumerate(selected_case_ids[:top_k], start=1):
        rec = corpus_rows[case_idx_by_id[cid]]
        idx = case_idx_by_id[cid]
        ranked_results.append({
            "rank": rank,
            "case_id": cid,
            "distance": float(distances[idx]),
            "case_score": float(case_scores[idx]),
            "query_score": q_score,
            "fact_value": float(rec["severity_scores"]["Fact"]),
            "injury_value": float(rec["severity_scores"]["Injury"]),
            "comp_value": float(rec["severity_scores"]["Compensation"]),
            "fact_text": rec.get("LI_fact") or rec.get("fact_text") or "",
            "injury_text": rec.get("LI_injury") or "",
            "comp_text": rec.get("LI_compensation") or rec.get("compensation_text") or "",
            "conclusion_text": rec.get("conclusion_text") or "",
            "case_full_text": build_case_full_text(rec),
            "retrieval_side": source_side,
            "anchor_case_id": anchor_case_id,
        })
    metadata = {
        "anchor_case_id": anchor_case_id,
        "anchor_distance": float(distances[anchor_idx]),
        "anchor_score": float(case_scores[anchor_idx]),
        "lighter_candidates": lighter_case_ids,
        "heavier_candidates": heavier_case_ids,
    }
    return ranked_results, metadata


def collect_descendant_cases(
    anchor_case_id: str,
    child_map: dict[str, list[str]],
    limit: int,
    distances: np.ndarray,
    case_idx_by_id: dict[str, int],
) -> list[str]:
    if limit <= 0:
        return []
    results: list[str] = []
    visited = set()
    current_level = list(child_map.get(anchor_case_id, []))
    while current_level and len(results) < limit:
        ordered_level = sorted(
            [cid for cid in current_level if cid not in visited],
            key=lambda cid: distances[case_idx_by_id[cid]],
        )
        next_level: list[str] = []
        for cid in ordered_level:
            if cid in visited:
                continue
            visited.add(cid)
            results.append(cid)
            if len(results) >= limit:
                break
            next_level.extend(child_map.get(cid, []))
        current_level = next_level
    return results


def ordered_boolean_matrix_for_ui(boolean_matrix: dict, legal_keys: dict[str, list[str]]) -> dict[str, list[dict]]:
    group_keys = {
        "litigants": ["single_plaintiff", "multiple_plaintiffs", "single_defendant", "multiple_defendants"],
        "fact": legal_keys.get("fact", []),
        "injury": legal_keys.get("injury", []),
        "compensation": legal_keys.get("compensation", []),
    }
    result: dict[str, list[dict]] = {}
    for group, keys in group_keys.items():
        values = boolean_matrix.get(group, {})
        result[group] = [
            {"key": key, "value": int(float(values.get(key, 0)))}
            for key in keys
        ]
    return result


def feature_values(boolean_matrix: dict, group: str, keys: list[str]) -> list[int]:
    values = boolean_matrix.get(group, {})
    return [int(float(values.get(key, 0))) for key in keys]


def overlap_stats(query_matrix: dict, case_matrix: dict, legal_keys: dict[str, list[str]]) -> dict[str, dict]:
    stats: dict[str, dict] = {}
    for group in ("fact", "injury", "compensation"):
        keys = legal_keys.get(group, [])
        q_values = feature_values(query_matrix, group, keys)
        c_values = feature_values(case_matrix, group, keys)
        shared = [
            key
            for key, q_value, c_value in zip(keys, q_values, c_values)
            if q_value == 1 and c_value == 1
        ]
        union_count = sum(1 for q_value, c_value in zip(q_values, c_values) if q_value == 1 or c_value == 1)
        shared_count = len(shared)
        normalized_distance = 0.0 if union_count == 0 else (union_count - shared_count) / union_count
        stats[group] = {
            "shared_activated_features": shared,
            "shared_count": shared_count,
            "activated_union_count": union_count,
            "normalized_binary_distance": float(normalized_distance),
        }
    return stats


def severity_direction(left_score: float, right_score: float, left_name: str, right_name: str) -> dict:
    if left_score < right_score:
        relation = "LH"
        description = f"{left_name} is lighter than {right_name}"
    elif left_score > right_score:
        relation = "HL"
        description = f"{left_name} is heavier than {right_name}"
    else:
        relation = "equal"
        description = f"{left_name} and {right_name} have equal weighted severity"
    return {
        "relation": relation,
        "description": description,
        "severity_difference": float(abs(left_score - right_score)),
    }


def build_query_case_relations(
    query_row: dict,
    exp: dict,
    similar_cases: list[dict],
    retrieval_meta: dict,
    corpus_by_id: dict[str, dict],
    legal_keys: dict[str, list[str]],
) -> tuple[dict, list[dict]]:
    query_matrix = build_query_boolean_matrix(query_row["query_text"], query_row["category"])
    query_levels, query_scores = classify_query_levels(query_row["query_text"], query_matrix)
    fact_w = float(exp["fact_w"])
    injury_w = float(exp["injury_w"])
    comp_w = float(exp["comp_w"])
    query_score = (
        fact_w * float(query_scores["Fact"])
        + injury_w * float(query_scores["Injury"])
        + comp_w * float(query_scores["Compensation"])
    )
    query_profile = {
        "severity_levels": query_levels,
        "severity_scores": query_scores,
        "weighted_severity_score": float(query_score),
        "boolean_matrix": ordered_boolean_matrix_for_ui(query_matrix, legal_keys),
    }
    anchor_case_id = str(retrieval_meta.get("anchor_case_id", ""))
    anchor_score = float(retrieval_meta.get("anchor_score", 0.0))

    relations = []
    for case in similar_cases:
        case_id = str(case["case_id"])
        rec = corpus_by_id.get(case_id, {})
        case_matrix = rec.get("boolean_matrix", {})
        case_score = float(case["case_score"])
        stats = overlap_stats(query_matrix, case_matrix, legal_keys)
        shared_any = any(group_stats["shared_count"] > 0 for group_stats in stats.values())
        relations.append({
            "rank": case["rank"],
            "case_id": case_id,
            "retrieval_source": case.get("retrieval_side", ""),
            "anchor_case_id": anchor_case_id,
            "is_anchor": case_id == anchor_case_id,
            "combined_distance_to_query": float(case["distance"]),
            "query_weighted_severity_score": float(query_score),
            "case_weighted_severity_score": case_score,
            "query_case_severity_relation": severity_direction(query_score, case_score, "query", "case"),
            "anchor_case_severity_relation": severity_direction(anchor_score, case_score, "anchor", "case"),
            "case_severity_scores": {
                "Fact": float(case["fact_value"]),
                "Injury": float(case["injury_value"]),
                "Compensation": float(case["comp_value"]),
            },
            "query_case_overlap": stats,
            "has_shared_activated_feature": shared_any,
            "case_boolean_matrix": ordered_boolean_matrix_for_ui(case_matrix, legal_keys),
            "fact_summary": case.get("fact_text", ""),
            "injury_summary": case.get("injury_text", ""),
            "compensation_summary": case.get("comp_text", ""),
            "case_full_text": case.get("case_full_text") or build_case_full_text(rec),
        })
    return query_profile, relations


def determine_applicable_laws(accident_facts: str, injuries: str, comp_facts: str, parties: dict) -> list[str]:
    return determine_applicable_laws_structured(accident_facts, injuries, comp_facts, parties)


def law_descriptions() -> dict[str, str]:
    return {
        "民法第184條第1項前段": "因故意或過失，不法侵害他人之權利者，負損害賠償責任。",
        "民法第185條第1項": "數人共同不法侵害他人之權利者，連帶負損害賠償責任。",
        "民法第187條第1項": "無行為能力人或限制行為能力人，不法侵害他人之權利者，以行為時有識別能力為限，與其法定代理人連帶負損害賠償責任。",
        "民法第188條第1項本文": "受僱人因執行職務，不法侵害他人之權利者，由僱用人與行為人連帶負損害賠償責任。",
        "民法第190條第1項": "動物加損害於他人者，由其占有人負損害賠償責任。",
        "民法第191條之2": "汽車、機車或其他非依軌道行駛之動力車輛，在使用中加損害於他人者，駕駛人應賠償因此所生之損害。",
        "民法第193條第1項": "不法侵害他人之身體或健康者，對於被害人因此喪失或減少勞動能力或增加生活上之需要時，應負損害賠償責任。",
        "民法第195條第1項前段": "不法侵害他人之身體、健康、名譽、自由、信用、隱私、貞操，或不法侵害其他人格法益而情節重大者，被害人雖非財產上之損害，亦得請求賠償相當之金額。",
    }


def build_reasoning_context(similar_cases: list[dict], corpus_by_id: dict[str, dict], parent_map: dict[str, str]) -> str:
    blocks = []
    for case in similar_cases:
        parent_id = parent_map.get(case["case_id"])
        parent_summary = ""
        if parent_id and parent_id in corpus_by_id:
            parent_rec = corpus_by_id[parent_id]
            parent_summary = (
                f"上位父案例 {parent_id}："
                f"F={parent_rec['severity_scores']['Fact']}, "
                f"I={parent_rec['severity_scores']['Injury']}, "
                f"C={parent_rec['severity_scores']['Compensation']}；"
                f"結論摘要={parent_rec.get('conclusion_text','')[:120]}"
            )
        blocks.append(
            f"案例 {case['case_id']}（distance={case['distance']:.4f}, score={case['case_score']:.4f}）"
            f"\n- 事故：{case['fact_text'][:240]}"
            f"\n- 受傷：{case['injury_text'][:180]}"
            f"\n- 賠償：{case['comp_text'][:240]}"
            + (f"\n- 圖譜推理：{parent_summary}" if parent_summary else "")
        )
    return "\n\n".join(blocks)


def call_llm(prompt: str, model: str, timeout: int = 180) -> str:
    response = requests.post(
        LLM_URL,
        json={"model": model, "prompt": prompt, "stream": False},
        timeout=timeout,
    )
    response.raise_for_status()
    return response.json()["response"].strip()


def generate_standard_facts(accident_facts: str, model: str) -> str:
    prompt = build_strict_facts_prompt(accident_facts)
    result = call_llm(prompt, model)
    result = re.sub(r"^一、\s*\n+\s*", "一、", result.strip())
    m = re.search(r"一、\s*(.*)", result, re.S)
    if m:
        body = m.group(1).strip()
        body = re.sub(r"^緣\s*\n+\s*", "緣", body)
        return f"一、{body}"
    text = re.sub(r'^一[、．.\s]*事故發生緣由[:：]?\s*', '', accident_facts.strip())
    return f"一、緣{text}"


def generate_standard_laws(accident_facts: str, injuries: str, comp_facts: str, parties: dict) -> str:
    laws = determine_applicable_laws(accident_facts, injuries, comp_facts, parties)
    desc = law_descriptions()
    law_texts = [f"「{desc[law]}」" for law in laws if law in desc]
    joined = "、".join(law_texts)
    joined_articles = "、".join(laws)
    return f"二、按{joined}，{joined_articles}分別定有明文。查被告因上開侵權行為，致原告受有下列損害，依前揭規定，被告應負損害賠償責任："


def topk_style_level(top_k: int) -> int:
    if top_k <= 1:
        return 0
    if top_k <= 3:
        return 1
    if top_k <= 5:
        return 2
    if top_k <= 8:
        return 3
    if top_k == 9:
        return 4
    return 5


def generate_compensation(comp_facts: str, injuries: str, accident_facts: str, parties: dict, model: str, style_level: int = 0) -> str:
    constraints = extract_damage_constraints(comp_facts, injuries)
    return build_structured_damage_section(comp_facts, injuries, parties, constraints, accident_facts, style_level)


def generate_conclusion(accident_facts: str, comp_facts: str, damage_section: str, parties: dict, model: str, style_level: int = 0) -> str:
    return build_conclusion_section(damage_section, parties, style_level)


def assemble_final_lawsuit(facts: str, laws: str, damages: str, conclusion: str) -> str:
    return cleanup_generated_lawsuit_text(f"{facts}\n\n{laws}\n\n{damages}\n\n{conclusion}".strip())


def cleanup_generated_lawsuit_text(text: str) -> str:
    text = text.replace("被告被告", "被告")
    text = text.replace("原告機車", "系爭機車")
    text = text.replace("原告汽車", "系爭汽車")
    text = re.sub(r"並有並且有", "並有", text)
    text = re.sub(r"有並且有", "並有", text)
    text = re.sub(r"並有[、，；\s]+", "並有", text)
    text = re.sub(r"原告(?:並且|且)原告", "原告", text)
    damage_labels = (
        "醫療費用|醫藥費用|醫藥費|醫療復健費用|交通費用|交通費|看護費用|看護費|"
        "工作損失|不能工作之損失|無法工作損失|薪資損失|收入損失|"
        "勞動能力減損|勞動能力損失|車輛修復費用|車輛修理費|財物損失|"
        "其他必要費用|精神慰撫金|慰撫金"
    )
    name_chars = r"\u3400-\u9fff\uf900-\ufaff○"
    name = rf"(?:原告)?[{name_chars}]{{0,5}}"
    amount = r"(?:[0-9][0-9,]*(?:萬[0-9,]*)?\s*元)?"
    text = re.sub(rf"(?m)^(\s*){name}(?:之)?(?:{damage_labels})\s*{amount}\s*[:：]\s*", r"\1", text)
    text = re.sub(rf"(?<=[。；\n])\s*{name}(?:之)?(?:{damage_labels})\s*{amount}\s*[:：]\s*", "", text)
    text = re.sub(rf"(?<=[，,])\s*(?:{damage_labels})\s*{amount}\s*[:：]\s*", "", text)
    text = balance_parentheses(text)
    return text


def balance_parentheses(text: str) -> str:
    cleaned_lines = []
    for line in text.splitlines():
        result = []
        ascii_open = 0
        full_open = 0
        for ch in line:
            if ch == "(":
                ascii_open += 1
                result.append(ch)
            elif ch == ")":
                if ascii_open > 0:
                    ascii_open -= 1
                    result.append(ch)
            elif ch == "（":
                full_open += 1
                result.append(ch)
            elif ch == "）":
                if full_open > 0:
                    full_open -= 1
                    result.append(ch)
            else:
                result.append(ch)
        cleaned = "".join(result)
        missing = ("）" * full_open) + (")" * ascii_open)
        if missing:
            if cleaned.endswith(("。", "；", "，")):
                cleaned = cleaned[:-1] + missing + cleaned[-1]
            else:
                cleaned += missing
        cleaned_lines.append(cleaned)
    return "\n".join(cleaned_lines)


def run_generation_for_query(
    query_row: dict,
    exp: dict,
    corpus_rows: list[dict],
    litigant_values: np.ndarray,
    legal_features: dict[str, np.ndarray],
    legal_keys: dict[str, list[str]],
    fact_values: np.ndarray,
    injury_values: np.ndarray,
    comp_values: np.ndarray,
    case_sort: np.ndarray,
    corpus_by_id: dict[str, dict],
    case_idx_by_id: dict[str, int],
    top_k: int,
    model: str,
) -> dict:
    tree_context = get_experiment_tree_context(
        exp,
        [str(rec["case_id"]) for rec in corpus_rows],
        case_sort,
        litigant_values,
        legal_features,
        fact_values,
        injury_values,
        comp_values,
    )
    similar_cases, retrieval_meta = retrieve_dual_tree_cases(
        query_row,
        exp,
        corpus_rows,
        litigant_values,
        legal_features,
        legal_keys,
        fact_values,
        injury_values,
        comp_values,
        case_sort,
        top_k,
        tree_context,
        case_idx_by_id,
    )
    sections = extract_sections(query_row["query_text"])
    parties = extract_parties(query_row["query_text"], model)
    reasoning_context = build_reasoning_context(similar_cases, corpus_by_id, tree_context.heavy_parent_map)
    generation_support = build_generation_support_context(similar_cases, tree_context.heavy_parent_map, corpus_by_id)
    query_feature_profile, query_case_relations = build_query_case_relations(
        query_row,
        exp,
        similar_cases,
        retrieval_meta,
        corpus_by_id,
        legal_keys,
    )
    style_level = topk_style_level(top_k)
    facts = generate_standard_facts(sections["accident_facts"] or query_row["query_text"], model)
    facts = cleanup_generated_lawsuit_text(facts)
    laws = generate_standard_laws(
        sections["accident_facts"] or query_row["query_text"],
        sections["injuries"],
        sections["compensation_facts"],
        parties,
    )
    laws = cleanup_generated_lawsuit_text(laws)
    damages = generate_compensation(
        sections["compensation_facts"] or query_row["query_text"],
        sections["injuries"],
        sections["accident_facts"],
        parties,
        model,
        style_level,
    )
    damages = cleanup_generated_lawsuit_text(damages)
    conclusion = generate_conclusion(
        sections["accident_facts"] or query_row["query_text"],
        sections["compensation_facts"],
        damages,
        parties,
        model,
        style_level,
    )
    conclusion = cleanup_generated_lawsuit_text(conclusion)
    generated_text = assemble_final_lawsuit(facts, laws, damages, conclusion)
    return {
        "query_id": query_row["query_id"],
        "category": query_row["category"],
        "exp_id": exp["exp_id"],
        "exp_name": exp["short_name"],
        "weight_code": exp["weight_code"],
        "tau_code": exp["tau_code"],
        "fact_w": exp["fact_w"],
        "injury_w": exp["injury_w"],
        "comp_w": exp["comp_w"],
        "tau": exp["distance_threshold"],
        "top_k": top_k,
        "topk_style_level": style_level,
        "model": model,
        "parties": parties,
        "retrieval_mode": "dual_tree",
        "tree_exp_id": tree_context.exp_id,
        "anchor_case_id": retrieval_meta["anchor_case_id"],
        "anchor_distance": retrieval_meta["anchor_distance"],
        "anchor_score": retrieval_meta["anchor_score"],
        "lighter_candidates": retrieval_meta["lighter_candidates"],
        "heavier_candidates": retrieval_meta["heavier_candidates"],
        "query_text": query_row["query_text"],
        "silver_reference_text": query_row["ground_truth_text"],
        "legacy_silver_reference_text": query_row["ground_truth_text"],
        "reference_text": query_row["ground_truth_text"],
        "human_reference_text": query_row["ground_truth_text"],
        "ground_truth_text": query_row["ground_truth_text"],
        "gpt_baseline_text": query_row["gpt_baseline_text"],
        "similar_cases": similar_cases,
        "query_feature_profile": query_feature_profile,
        "query_case_relations": query_case_relations,
        "parent_map": tree_context.heavy_parent_map,
        "reverse_parent_map": tree_context.light_parent_map,
        "reasoning_context": reasoning_context,
        "generation_support": generation_support,
        "facts_section": facts,
        "laws_section": laws,
        "damages_section": damages,
        "conclusion_section": conclusion,
        "generated_text": generated_text,
    }


def save_batch_outputs(records: list[dict], output_stem: str) -> tuple[Path, Path]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    jsonl_path = OUTPUT_DIR / f"{output_stem}.jsonl"
    csv_path = OUTPUT_DIR / f"{output_stem}.csv"

    with jsonl_path.open("w", encoding="utf-8") as fh:
        for record in records:
            fh.write(json.dumps(record, ensure_ascii=False) + "\n")

    fieldnames = csv_fieldnames(include_status=False)
    with csv_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(csv_record(record, fieldnames))

    return jsonl_path, csv_path


def csv_fieldnames(include_status: bool) -> list[str]:
    fieldnames = [
        "query_id",
        "category",
        "exp_id",
        "exp_name",
        "retrieval_mode",
        "tree_exp_id",
        "anchor_case_id",
        "anchor_distance",
        "anchor_score",
        "weight_code",
        "tau_code",
        "fact_w",
        "injury_w",
        "comp_w",
        "tau",
        "top_k",
        "topk_style_level",
        "model",
        "query_text",
        "silver_reference_text",
        "legacy_silver_reference_text",
        "human_reference_text",
        "ground_truth_text",
        "gpt_baseline_text",
        "top8_case_ids",
        "top8_distances",
        "top8_retrieval_sources",
        "top8_query_case_relations",
        "top8_case_full_texts",
        "query_feature_profile",
        "query_case_relations",
        "similar_cases",
        "generation_support",
        "facts_section",
        "laws_section",
        "damages_section",
        "conclusion_section",
        "generated_text",
    ]
    if include_status:
        fieldnames.extend(["run_status", "error_message"])
    return fieldnames


def csv_safe_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def csv_record(record: dict, fieldnames: list[str]) -> dict:
    row = {name: csv_safe_value(record.get(name, "")) for name in fieldnames}
    relations = record.get("query_case_relations") or []
    if relations:
        row["top8_case_ids"] = " | ".join(str(item.get("case_id", "")) for item in relations)
        row["top8_distances"] = " | ".join(f"{float(item.get('combined_distance_to_query', 0.0)):.4f}" for item in relations)
        row["top8_retrieval_sources"] = " | ".join(str(item.get("retrieval_source", "")) for item in relations)
        row["top8_query_case_relations"] = " | ".join(
            f"{item.get('case_id', '')}:{(item.get('query_case_severity_relation') or {}).get('relation', '')}"
            for item in relations
        )
        row["top8_case_full_texts"] = "\n\n===== SDKG TOP CASE =====\n\n".join(
            f"case_id={item.get('case_id', '')}\n{item.get('case_full_text', '')}"
            for item in relations
        )
    return row


def output_paths_for_stem(output_stem: str) -> tuple[Path, Path]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return OUTPUT_DIR / f"{output_stem}.jsonl", OUTPUT_DIR / f"{output_stem}.csv"


def append_record_jsonl(path: Path, record: dict) -> None:
    with path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(record, ensure_ascii=False) + "\n")


def rewrite_progress_csv(path: Path, records: list[dict]) -> None:
    fieldnames = csv_fieldnames(include_status=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(csv_record(record, fieldnames))


def select_target_experiments(exp_arg: str | None) -> list[dict]:
    experiments = build_experiments()
    if exp_arg is None:
        return experiments

    experiments_by_id = {exp["exp_id"]: exp for exp in experiments}
    experiments_by_name = {exp["short_name"]: exp for exp in experiments}
    exp = experiments_by_id.get(exp_arg) or experiments_by_name.get(exp_arg)
    if exp is None:
        known = ", ".join(list(experiments_by_id.keys()) + list(experiments_by_name.keys()))
        raise ValueError(f"Unknown exp_id {exp_arg}. Known values: {known}")
    return [exp]


def build_default_output_stem(exp: dict | None, top_k: int, all_queries: bool, all_experiments: bool, query_id: int | None) -> str:
    if all_experiments and all_queries:
        return f"sdkg_full_18exp_50q_topk{top_k}"
    if all_experiments:
        return f"sdkg_all_exp_query_{query_id}_topk{top_k}"
    if all_queries and exp is not None:
        return f"sdkg_batch_{exp['short_name']}_topk{top_k}"
    if exp is not None and query_id is not None:
        return f"sdkg_query_{query_id}_{exp['short_name']}_topk{top_k}"
    return f"sdkg_run_topk{top_k}"


def parse_query_ids_arg(query_ids_arg: str | None) -> list[int]:
    if not query_ids_arg:
        return []
    query_ids: list[int] = []
    for part in query_ids_arg.split(","):
        part = part.strip()
        if not part:
            continue
        query_ids.append(int(part))
    return query_ids


def main() -> None:
    global ACTIVE_LAMBDA_LEGAL

    parser = argparse.ArgumentParser()
    parser.add_argument("--query-id", type=int, default=1)
    parser.add_argument("--query-ids", default="", help="Comma-separated query_ids, e.g. 3,17,27")
    parser.add_argument("--exp-id", default="E05")
    parser.add_argument("--top-k", type=int, default=3)
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--save", action="store_true")
    parser.add_argument("--all-queries", action="store_true")
    parser.add_argument("--all-experiments", action="store_true")
    parser.add_argument("--output-stem", default="")
    parser.add_argument("--lambda-legal", type=float, default=ACTIVE_LAMBDA_LEGAL)
    args = parser.parse_args()

    ACTIVE_LAMBDA_LEGAL = float(args.lambda_legal)

    corpus_rows, litigant_values, legal_features, legal_keys, fact_values, injury_values, comp_values, case_sort = load_case_corpus()
    queries = load_queries(QUERY_FILE)
    target_experiments = select_target_experiments(None if args.all_experiments else args.exp_id)
    corpus_by_id = {str(rec["case_id"]): rec for rec in corpus_rows}
    case_idx_by_id = {str(rec["case_id"]): idx for idx, rec in enumerate(corpus_rows)}
    requested_query_ids = parse_query_ids_arg(args.query_ids)
    if args.all_queries:
        target_queries = queries
    elif requested_query_ids:
        query_by_id = {row["query_id"]: row for row in queries}
        missing = [qid for qid in requested_query_ids if qid not in query_by_id]
        if missing:
            raise ValueError(f"query_ids {missing} not found in {QUERY_FILE}")
        target_queries = [query_by_id[qid] for qid in requested_query_ids]
    else:
        query_row = next((row for row in queries if row["query_id"] == args.query_id), None)
        if query_row is None:
            raise ValueError(f"query_id {args.query_id} not found in {QUERY_FILE}")
        target_queries = [query_row]

    records = []
    verbose_single_run = (len(target_experiments) == 1 and len(target_queries) == 1)
    total_jobs = len(target_experiments) * len(target_queries)
    job_idx = 0

    output_stem = ""
    jsonl_path = None
    csv_path = None
    if args.save:
        if args.output_stem:
            output_stem = args.output_stem
        else:
            exp_for_name = target_experiments[0] if len(target_experiments) == 1 else None
            query_id_for_name = target_queries[0]["query_id"] if len(target_queries) == 1 else None
            output_stem = build_default_output_stem(
                exp_for_name,
                args.top_k,
                args.all_queries,
                args.all_experiments,
                query_id_for_name,
            )
        jsonl_path, csv_path = output_paths_for_stem(output_stem)
        jsonl_path.write_text("", encoding="utf-8")
        rewrite_progress_csv(csv_path, [])

    for exp in target_experiments:
        for query_row in target_queries:
            job_idx += 1
            print(
                f"[{job_idx}/{total_jobs}] running query_id={query_row['query_id']} "
                f"exp={exp['short_name']} top_k={args.top_k} model={args.model}",
                flush=True,
            )
            try:
                record = run_generation_for_query(
                    query_row,
                    exp,
                    corpus_rows,
                    litigant_values,
                    legal_features,
                    legal_keys,
                    fact_values,
                    injury_values,
                    comp_values,
                    case_sort,
                    corpus_by_id,
                    case_idx_by_id,
                    args.top_k,
                    args.model,
                )
                record["run_status"] = "ok"
                record["error_message"] = ""
                print(
                    f"[{job_idx}/{total_jobs}] done query_id={query_row['query_id']} exp={exp['short_name']}",
                    flush=True,
                )
            except Exception as exc:
                record = {
                    "query_id": query_row["query_id"],
                    "category": query_row["category"],
                    "exp_id": exp["exp_id"],
                    "exp_name": exp["short_name"],
                    "retrieval_mode": "dual_tree",
                    "tree_exp_id": exp["exp_id"],
                    "anchor_case_id": "",
                    "anchor_distance": "",
                    "anchor_score": "",
                    "weight_code": exp["weight_code"],
                    "tau_code": exp["tau_code"],
                    "fact_w": exp["fact_w"],
                    "injury_w": exp["injury_w"],
                    "comp_w": exp["comp_w"],
                    "tau": exp["distance_threshold"],
                    "top_k": args.top_k,
                    "model": args.model,
                    "query_text": query_row["query_text"],
                    "silver_reference_text": query_row["ground_truth_text"],
                    "legacy_silver_reference_text": query_row["ground_truth_text"],
                    "human_reference_text": query_row["ground_truth_text"],
                    "ground_truth_text": query_row["ground_truth_text"],
                    "gpt_baseline_text": query_row["gpt_baseline_text"],
                    "facts_section": "",
                    "laws_section": "",
                    "damages_section": "",
                    "conclusion_section": "",
                    "generated_text": "",
                    "run_status": "error",
                    "error_message": str(exc),
                }
                print(
                    f"[{job_idx}/{total_jobs}] failed query_id={query_row['query_id']} exp={exp['short_name']} error={exc}",
                    flush=True,
                )

            records.append(record)
            if args.save and jsonl_path is not None and csv_path is not None:
                append_record_jsonl(jsonl_path, record)
                rewrite_progress_csv(csv_path, records)

            if verbose_single_run:
                print("=" * 100)
                print(
                    f"SDKG query_id={record['query_id']} exp_id={record['exp_id']} "
                    f"exp_name={record['exp_name']} top_k={args.top_k} model={args.model}"
                )
                print(f"category={record['category']}")
                print(f"parties={record['parties']}")
                print("-" * 100)
                print("【找到的相似案例】")
                for case in record["similar_cases"]:
                    print(f"#{case['rank']} case_id={case['case_id']} distance={case['distance']:.4f} score={case['case_score']:.4f}")
                    print(f"事故事實：{case['fact_text'][:260]}")
                    print(f"受傷情形：{case['injury_text'][:220]}")
                    print(f"賠償依據：{case['comp_text'][:260]}")
                    parent_id = record["parent_map"].get(case["case_id"])
                    if parent_id:
                        print(f"圖譜推理：parent_case_id={parent_id}")
                    print("-" * 100)
                print("【圖譜推理摘要】")
                print(record["reasoning_context"])
                print("-" * 100)
                print("【生成段落：事實】")
                print(record["facts_section"])
                print("-" * 100)
                print("【生成段落：法條】")
                print(record["laws_section"])
                print("-" * 100)
                print("【生成段落：損害】")
                print(record["damages_section"])
                print("-" * 100)
                print("【生成段落：結論】")
                print(record["conclusion_section"])
                print("-" * 100)
                print("【生成起訴書】")
                print(record["generated_text"])
                print("=" * 100)

    if args.save and jsonl_path is not None and csv_path is not None:
        print(f"saved_jsonl={jsonl_path}")
        print(f"saved_csv={csv_path}")


if __name__ == "__main__":
    main()
